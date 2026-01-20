from flask import Flask, render_template, request, redirect, session, flash, send_file
import sqlite3, random, string, os, io
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from datetime import datetime
import qrcode
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm

# ================= APP CONFIG =================
app = Flask(__name__)
app.secret_key = "fake_product_detection_secure"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE_DIR, "database.db")
EXCEL = os.path.join(BASE_DIR, "generated_product_codes.xlsx")
QR_DIR = os.path.join(BASE_DIR, "qr_codes")
os.makedirs(QR_DIR, exist_ok=True)

# ================= DATABASE =================
def get_db():
    con = sqlite3.connect(DB, timeout=30, check_same_thread=False)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    with get_db() as con:
        cur = con.cursor()

        cur.execute("""
        CREATE TABLE IF NOT EXISTS user (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            role TEXT
        )""")

        cur.execute("""
        CREATE TABLE IF NOT EXISTS product (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            brand TEXT,
            manufactured DATE,
            status TEXT DEFAULT 'ACTIVE'
        )""")

        cur.execute("""
        CREATE TABLE IF NOT EXISTS product_code (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            code TEXT UNIQUE
        )""")

        cur.execute("""
        CREATE TABLE IF NOT EXISTS complaint_token (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_name TEXT,
            brand TEXT,
            code TEXT,
            issue TEXT,
            customer_name TEXT,
            customer_contact TEXT,
            admin_reply TEXT,
            status TEXT DEFAULT 'OPEN',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )""")

        # ensure status column exists for old DBs
        try:
            cur.execute("ALTER TABLE product ADD COLUMN status TEXT DEFAULT 'ACTIVE'")
        except sqlite3.OperationalError:
            pass

        users = [
            ("admin", "admin123", "admin"),
            ("worker", "worker123", "worker"),
            ("client", "client123", "client")
        ]

        for u in users:
            cur.execute(
                "INSERT OR IGNORE INTO user VALUES (NULL,?,?,?)",
                (u[0], generate_password_hash(u[1]), u[2])
            )

with app.app_context():
    init_db()

# ================= EXCEL =================
def ensure_excel():
    if not os.path.exists(EXCEL):
        wb = Workbook()
        ws = wb.active
        ws.append(["Product", "Brand", "Code", "Generated At"])
        wb.save(EXCEL)

def save_excel(name, brand, code):
    ensure_excel()
    wb = load_workbook(EXCEL)
    ws = wb.active
    ws.append([name, brand, code, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(EXCEL)

# ================= UTIL =================
def generate_code():
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=12))

def generate_qr(code):
    qrcode.make(code).save(os.path.join(QR_DIR, f"{code}.png"))

# ================= LOGIN =================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        with get_db() as con:
            user = con.execute(
                "SELECT * FROM user WHERE username=?",
                (request.form["username"],)
            ).fetchone()

        if user and check_password_hash(user["password_hash"], request.form["password"]):
            session["role"] = user["role"]
            return redirect(f"/{user['role']}")

        flash("Invalid username or password")

    return render_template("login.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ================= ADMIN =================
@app.route("/admin")
def admin():
    if session.get("role") != "admin":
        return redirect("/")

    search = request.args.get("search", "")

    with get_db() as con:
        products = con.execute("""
            SELECT * FROM product
            WHERE status='ACTIVE'
            AND (name LIKE ? OR brand LIKE ?)
            ORDER BY id DESC
        """, (f"%{search}%", f"%{search}%")).fetchall()

        archived = con.execute(
            "SELECT * FROM product WHERE status='ARCHIVED'"
        ).fetchall()

        tokens = con.execute(
            "SELECT * FROM complaint_token ORDER BY created_at DESC"
        ).fetchall()

    return render_template(
        "admin_dashboard.html",
        products=products,
        archived=archived,
        tokens=tokens,
        search=search
    )

@app.route("/admin/add_product", methods=["POST"])
def add_product():
    with get_db() as con:
        con.execute(
            "INSERT INTO product (name,brand,manufactured) VALUES (?,?,?)",
            (request.form["name"], request.form["brand"], request.form["manufactured"])
        )
    return redirect("/admin")

@app.route("/admin/generate/<int:pid>", methods=["POST"])
def generate(pid):
    qty = int(request.form["qty"])
    with get_db() as con:
        p = con.execute("SELECT * FROM product WHERE id=?", (pid,)).fetchone()
        for _ in range(qty):
            code = generate_code()
            con.execute(
                "INSERT INTO product_code (product_id,code) VALUES (?,?)",
                (pid, code)
            )
            save_excel(p["name"], p["brand"], code)
            generate_qr(code)
    return redirect("/admin")

# -------- DATA CONTROL --------

@app.route("/admin/archive_product/<int:pid>", methods=["POST"])
def archive_product(pid):
    with get_db() as con:
        con.execute(
            "UPDATE product SET status='ARCHIVED' WHERE id=?", (pid,)
        )
    flash("Product archived")
    return redirect("/admin")

@app.route("/admin/delete_product/<int:pid>", methods=["POST"])
def delete_product(pid):
    with get_db() as con:
        con.execute("DELETE FROM product_code WHERE product_id=?", (pid,))
        con.execute("DELETE FROM product WHERE id=?", (pid,))
    flash("Product permanently deleted")
    return redirect("/admin")

@app.route("/admin/delete_code/<int:cid>", methods=["POST"])
def delete_code(cid):
    with get_db() as con:
        con.execute("DELETE FROM product_code WHERE id=?", (cid,))
    flash("QR code deleted")
    return redirect("/admin")

# ================= DOWNLOADS =================
@app.route("/admin/download_excel")
def download_excel():
    ensure_excel()
    return send_file(EXCEL, as_attachment=True)

@app.route("/admin/download_qr_pdf/<int:product_id>")
def download_qr_pdf(product_id):
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    with get_db() as con:
        product = con.execute(
            "SELECT * FROM product WHERE id=?", (product_id,)
        ).fetchone()
        codes = con.execute(
            "SELECT code FROM product_code WHERE product_id=?", (product_id,)
        ).fetchall()

    x, y = 2 * cm, height - 3 * cm
    for c in codes:
        path = os.path.join(QR_DIR, f"{c['code']}.png")
        if os.path.exists(path):
            pdf.drawImage(path, x, y, 4 * cm, 4 * cm)
            pdf.drawString(x + 5 * cm, y + 2 * cm, product["name"])
            pdf.drawString(x + 5 * cm, y + 1.2 * cm, c["code"])
            y -= 5 * cm
            if y < 5 * cm:
                pdf.showPage()
                y = height - 3 * cm

    pdf.save()
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{product['name']}_QR.pdf",
        mimetype="application/pdf"
    )

# ================= WORKER =================
@app.route("/worker", methods=["GET", "POST"])
def worker():
    if session.get("role") != "worker":
        return redirect("/")
    result = None
    if request.method == "POST":
        code = request.form["code"]
        with get_db() as con:
            row = con.execute("SELECT * FROM product_code WHERE code=?", (code,)).fetchone()
        result = "GENUINE" if row else "FAKE"
    return render_template("worker_dashboard.html", result=result)

# ================= CLIENT =================
@app.route("/client", methods=["GET", "POST"])
def client():
    if session.get("role") != "client":
        return redirect("/")

    result = None
    if request.method == "POST":
        code = request.form["code"]
        with get_db() as con:
            row = con.execute("""
                SELECT p.name,p.brand,p.manufactured
                FROM product_code pc
                JOIN product p ON pc.product_id=p.id
                WHERE pc.code=? AND p.status='ACTIVE'
            """, (code,)).fetchone()

        if row:
            result = {
                "status": "GENUINE",
                "code": code,
                "name": row["name"],
                "brand": row["brand"],
                "manufactured": row["manufactured"]
            }
        else:
            result = {"status": "FAKE", "code": code}

    return render_template("client_dashboard.html", result=result)

@app.route("/client/raise_token", methods=["GET", "POST"])
def raise_token():
    if session.get("role") != "client":
        return redirect("/")

    code = request.args.get("code")
    if request.method == "POST":
        with get_db() as con:
            con.execute("""
                INSERT INTO complaint_token
                (product_name,brand,code,issue,customer_name,customer_contact)
                VALUES (?,?,?,?,?,?)
            """, (
                request.form["product_name"],
                request.form["brand"],
                request.form["code"],
                request.form["issue"],
                request.form["customer_name"],
                request.form["customer_contact"]
            ))
        flash("Complaint submitted")
        return redirect("/client")

    return render_template("client_tokens.html", code=code)

# ================= RUN =================
if __name__ == "__main__":
    app.run()
